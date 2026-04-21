"""
TaskTree 导入导出路由
====================
支持将项目数据导出为 JSON / Markdown / Excel 格式，
以及从 JSON 文件导入任务树和标签。
"""
import io
import json
from datetime import datetime, timezone
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.core.database import get_db
from app.models import User, Project, Task, TaskTag, TaskTagRelation, ProjectMember
from app.schemas import MessageResponse
from app.api.v1.auth import get_current_user
from sqlalchemy import and_

router = APIRouter()


async def _get_project_access(project_id: int, db: AsyncSession, current_user: User) -> Project:
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    if project.owner_id != current_user.id:
        result = await db.execute(
            select(ProjectMember).where(
                and_(ProjectMember.project_id == project_id, ProjectMember.user_id == current_user.id)
            )
        )
        if not result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="没有权限")
    return project


def _build_export_tree(tasks: list, parent_id=None) -> list:
    """构建导出用的任务树（O(n) 复杂度，使用 HashMap 预处理）。"""
    # 预建 parent_id -> children 映射
    children_map: dict = {}
    for t in tasks:
        pid = t.parent_id
        if pid not in children_map:
            children_map[pid] = []
        children_map[pid].append(t)

    def build(pid):
        result = []
        for t in children_map.get(pid, []):
            node = {
                "name": t.name,
                "description": t.description,
                "status": t.status,
                "priority": t.priority,
                "progress": t.progress,
                "start_date": str(t.start_date) if t.start_date else None,
                "due_date": str(t.due_date) if t.due_date else None,
                "estimated_time": t.estimated_time,
                "sort_order": t.sort_order,
                "children": build(t.id),
            }
            result.append(node)
        return result

    return build(parent_id)


# ========== JSON 导出 ==========
@router.get("/{project_id}/export/json")
async def export_json(
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await _get_project_access(project_id, db, current_user)

    # 获取任务
    result = await db.execute(select(Task).where(Task.project_id == project_id).order_by(Task.sort_order))
    tasks = result.scalars().all()

    # 获取标签
    result = await db.execute(select(TaskTag).where(TaskTag.project_id == project_id))
    tags = result.scalars().all()

    data = {
        "project_name": project.name,
        "project_description": project.description,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "tasks": _build_export_tree(list(tasks)),
        "tags": [{"name": t.name, "color": t.color} for t in tags],
    }

    content = json.dumps(data, ensure_ascii=False, indent=2)
    filename = quote(f"{project.name}.json")
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ========== Markdown 导出 ==========
@router.get("/{project_id}/export/markdown")
async def export_markdown(
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await _get_project_access(project_id, db, current_user)

    result = await db.execute(select(Task).where(Task.project_id == project_id).order_by(Task.sort_order))
    tasks = result.scalars().all()

    status_emoji = {"pending": "⬜", "in_progress": "🔄", "completed": "✅", "cancelled": "❌"}
    priority_emoji = {"high": "🔴", "medium": "🟡", "low": "🟢"}

    lines = [f"# {project.name}\n"]
    if project.description:
        lines.append(f"{project.description}\n")
    lines.append(f"导出时间: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}\n\n---\n")

    def render_tasks(task_list, level=0):
        for t in task_list:
            if t.parent_id is None or level > 0:
                indent = "  " * level
                check = "x" if t.status == "completed" else " "
                emoji = status_emoji.get(t.status, "⬜")
                pri = priority_emoji.get(t.priority, "")
                line = f"{indent}- [{check}] {emoji} {pri} **{t.name}**"
                if t.due_date:
                    line += f" (截止: {t.due_date})"
                if t.progress > 0 and t.status != "completed":
                    line += f" [{t.progress}%]"
                lines.append(line)
                if t.description:
                    lines.append(f"{indent}  > {t.description}")
                # 查找子任务
                children = [c for c in tasks if c.parent_id == t.id]
                if children:
                    render_tasks(children, level + 1)

    root_tasks = [t for t in tasks if t.parent_id is None]
    render_tasks(root_tasks)

    content = "\n".join(lines)
    filename = quote(f"{project.name}.md")
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/markdown",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ========== Excel 导出 ==========
@router.get("/{project_id}/export/excel")
async def export_excel(
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await _get_project_access(project_id, db, current_user)

    result = await db.execute(select(Task).where(Task.project_id == project_id).order_by(Task.sort_order))
    tasks = result.scalars().all()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="服务器缺少 openpyxl 依赖，无法导出 Excel")

    wb = Workbook()
    ws = wb.active
    ws.title = "任务列表"

    # 表头
    headers = ["ID", "任务名称", "状态", "优先级", "进度(%)", "开始日期", "截止日期", "预计工时(分)", "父任务ID"]
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_map = {"pending": "待办", "in_progress": "进行中", "completed": "已完成", "cancelled": "已取消"}
    priority_map = {"high": "高", "medium": "中", "low": "低"}

    for row, task in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=task.id)
        ws.cell(row=row, column=2, value=task.name)
        ws.cell(row=row, column=3, value=status_map.get(task.status, task.status))
        ws.cell(row=row, column=4, value=priority_map.get(task.priority, task.priority))
        ws.cell(row=row, column=5, value=task.progress)
        ws.cell(row=row, column=6, value=str(task.start_date) if task.start_date else "")
        ws.cell(row=row, column=7, value=str(task.due_date) if task.due_date else "")
        ws.cell(row=row, column=8, value=task.estimated_time or "")
        ws.cell(row=row, column=9, value=task.parent_id or "")

    # 自动列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{quote(f"{project.name}.xlsx")}"'},
    )


# ========== JSON 导入 ==========
@router.post("/{project_id}/import/json", response_model=MessageResponse)
async def import_json(
    project_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await _get_project_access(project_id, db, current_user)

    # 读取文件
    content = await file.read()
    try:
        data = json.loads(content.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        raise HTTPException(status_code=400, detail="无效的 JSON 文件")

    count = 0

    async def import_tasks(task_list: list, parent_id=None):
        nonlocal count
        for item in task_list:
            task = Task(
                project_id=project_id,
                parent_id=parent_id,
                name=item.get("name", "未命名任务"),
                description=item.get("description"),
                status=item.get("status", "pending"),
                priority=item.get("priority", "medium"),
                progress=item.get("progress", 0),
                start_date=item.get("start_date"),
                due_date=item.get("due_date"),
                estimated_time=item.get("estimated_time"),
                sort_order=item.get("sort_order", count),
            )
            db.add(task)
            await db.flush()
            count += 1

            children = item.get("children", [])
            if children:
                await import_tasks(children, task.id)

    # 导入任务
    tasks = data.get("tasks", [])
    if not tasks:
        raise HTTPException(status_code=400, detail="文件中没有任务数据")

    await import_tasks(tasks)

    # 导入标签
    tags = data.get("tags", [])
    for tag_data in tags:
        tag = TaskTag(
            project_id=project_id,
            name=tag_data.get("name", ""),
            color=tag_data.get("color"),
        )
        db.add(tag)

    await db.commit()

    return MessageResponse(
        code=201,
        message=f"成功导入 {count} 个任务",
        data={"task_count": count, "tag_count": len(tags)},
    )
